#! /usr/bin/env python3

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from config import configs

# 1. 加载xls文件
wb = load_workbook(filename="../xlsxFile/example.xlsx")
# 2. 获取表单
dailyTaskManager = wb['dailyTaskManager']
projectTaskManager = wb['projectTaskManager']
emailManager = wb['emailManager']
# 3. 对表单进行操作
# 3.0 获取表头信息
dailyTaskManagerTitle = []
for titleIndex in range(dailyTaskManager.min_column, dailyTaskManager.max_column+1):
    if dailyTaskManager.cell(1, titleIndex).value:
        dailyTaskManagerTitle.append( dailyTaskManager.cell(1, titleIndex).value)

for title in dailyTaskManagerTitle:
    print(title)
# 3.1 对表进行读操作
print(dailyTaskManager.min_column)
print(dailyTaskManager.max_column)
# 3.1 对表进行写操作
# 3.1 对表进行查操作
# 3.1 对表进行改操作

async def execute(tableOperation:Worksheet, **args):
    operationType = args.get('operationType', None)
    if operationType is None:
        # FIXME:操作错误，添加日志并返回
        return None

    if 'SAVE' == operationType:
        pass
    elif 'UPDATE' == operationType:
        pass
    elif 'DELETE' == operationType:
        pass
    elif 'FINDALL' == operationType:
        pass
    elif 'FIND' == operationType:
        pass
    else:
        # FIXME:操作错误，添加日志并返回
        pass
    

class ModelMetaclass(type):

    def __new__(cls, name, bases, attrs):
        print("##IN __new__")
        print("### IN ModelMetaclass __new__ name " , name)
        print("### IN ModelMetaclass __new__ attrs " , attrs)
        #FIXME: 读取配置文件加载文件
        
        global __filePath__
        __filePath__ = configs.xls.filePath
        global __fileName__
        __fileName__ = configs.xls.fileName

        if __filePath__ is None or __fileName__ is None:
            print('ERROR!!!\n')
            return None

        global __fileOperation__
        __fileOperation__ = load_workbook(filename=__filePath__ + __fileName__)

        if 'Model' == name:
            print('$$$$Model')
            return type.__new__(cls, name, bases, attrs)
        else:
            print("### OUT ModelMetaclass __new__ name " , name)
            print("### OUT ModelMetaclass __new__ bases.name " , bases)
            print("### OUT ModelMetaclass __new__ bases " , dir(bases))
            print("### OUT ModelMetaclass __new__ attrs " , attrs)
            tableName = attrs.get('__tableName__', None) or None
            #TODO: 对fileOperation添加类型判断
            if __fileOperation__ is None:
                #FIXME:  操作错误，添加日志并返回
                print("!!!!fileOperation in None!!!")
                return 
            
            tableOperation =  __fileOperation__[tableName]
            
            if tableOperation is None:
                #FIXME: 操作错误，添加日志并返回
                print("!!!!tableOperation in None!!!")
                return
            
            minLineIndex = tableOperation.min_row
            maxLineIndex = tableOperation.max_row
    
            maxColumIndex = tableOperation.max_column
            minColumIndex = tableOperation.min_column

            attrs['__tableOperation__'] = tableOperation
            attrs['__minLineIndex__'] = minLineIndex
            attrs['__maxLineIndex__'] = maxLineIndex
            attrs['__maxColumIndex__'] =  maxColumIndex
            attrs['__minColumIndex__'] = minColumIndex
        
        print("### OUT ModelMetaclass __new__ name " , name)
        print("### OUT ModelMetaclass __new__ bases " , dir(bases))
        print("### OUT ModelMetaclass __new__ attrs " , attrs)

        return type.__new__(cls, name, bases, attrs)

class Model(dict, metaclass=ModelMetaclass):
    __testXHH__ = 'test'
    def __init__(self, **kw):
        print("@@kw: ", kw)
        super(Model, self).__init__(**kw)

    def __getattr__(self, key):
        try:
            return self[key]
        except KeyError:
            raise AttributeError(r"'Model' object has no attribute '%s'" % key)
    
    def __setattr__(self, key, value):
        self[key] = value

    def getValue(self, key):
        return getattr(self, key, None)

    def getValueOrDefault(self, key):
        pass

    @classmethod
    async def findAll(cls, where=None, args=None, **kw):
        pass

    @classmethod
    async def find(cls, pkg):
        print('Model find!!!')

    @classmethod
    def save(self):
        print("XXXXXXXXXXXXXXX save!")
        print("__tableOperation__: ", self.__tableOperation__)
        __fileOperation__.save(filename=__filePath__ + __fileName__)


    def update(self, values):
        print("XXXXXXXXXXXXXXX update! table Name: ", self.__tableOperation__)
        insert_line = self.__maxLineIndex__
        if isinstance(values, list):
            for value in values:
                insert_line += 1
                if isinstance(value, list):
                    insert_column = self.__minColumIndex__
                    for v in value:
                        print("value: ", v)
                        print("insert_line: ", insert_line)
                        print("insert_column: ", insert_column)
                        ret = self.__tableOperation__.cell(row=insert_line, column=insert_column, value=v)
                        print("update ret: ", ret)
                        insert_column += 1
                else:
                    self.__maxLineIndex__ = insert_line + 1
                    print('error value')
                    return 

        else:
            self.__maxLineIndex__ = insert_line + 1
            print('error input values')
            return
        self.__maxLineIndex__ = insert_line + 1

    async def remove(self):
        pass


class DailyTaskManager(Model):
    __tableName__ = 'dailyTaskManager'
    def __init__(self):
        print("DailyTaskManager __init__")


class EmailManager(Model):
    __tableName__ = 'emailManager'
    def __init__(self):
        print("###beg emailManager __init__")
        dailyTaskManagerTitle = []
        for titleIndex in range(self.__minLineIndex__, self.__maxLineIndex__+1):
            for titleIndexCol in range(self.__minColumIndex__, self.__maxColumIndex__+1):
                if self.__tableOperation__.cell(titleIndex, titleIndexCol).value:
                    dailyTaskManagerTitle.append( self.__tableOperation__.cell(titleIndex, titleIndexCol).value)

        for title in dailyTaskManagerTitle:
            print(title)       
        print("###end emailManager __init__")

class TaskManager(Model):
    __tableName__ = 'projectTaskManager'
    def __init__(self):
        print("taskManager __init__")
    

if __name__ == "__main__":
    print("##test()##")
   # taskManager = TaskManager()
   # dailyTaskManager = DailyTaskManager()
    emailManager = EmailManager()
  #  taskManager.update([[1,2,3,4,5,6],[1,2,3,4,5,6],[1,2,3,4,5,6]])
   # dailyTaskManager.update([[1,2,3,4,5,6,7],[1,2,3,4,5,6,7],[1,2,3,4,5,6,7]])
    emailManager.update([[1,2,3,4],[1,2,3,4],[1,2,3,4]])
    emailManager.save()